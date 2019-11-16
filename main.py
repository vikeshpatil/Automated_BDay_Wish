#! python3
import datetime
import log
import send_sms
import send_mail
import whatsapp_msg
import fb_msg

import os
dir_path = os.path.dirname(os.path.realpath(__file__))

import csv
from openpyxl import load_workbook


today = datetime.date.today()
#------------- Reading data in text file----------------------#
def read_txt():

    with open(dir_path + '/credentials/text.txt', mode='r') as f:
        readContent = f.readlines()
    for j in range(len(readContent)):
        if(readContent[j].split()[4] == str(today)):
            Person = readContent[j].split()
            fname = Person[0]
            lname = Person[1]
            name = Person[0] + " " + Person[1]
            mobile = Person[2]
            email = Person[3]
            whatapp_contact_list = []  # contacts to send message
            for i in range(5, len(Person)):
                whatapp_name = Person[i].split('_')

                whatapp_name = '"' + ' '.join(whatapp_name) + '"'

                whatapp_contact_list.append(whatapp_name)

            print("Happy B'day " + name + " . Your mobile numbe is " + mobile + ". And your email is " + email )
            # send_sms.sendPostRequest(str(mobile), fname)     #send sms
            # send_mail.mail("rajp8340@gmail.com", fname)       #send mail
            whatsapp_msg.send_msg(whatapp_contact_list, fname)     #whatsapp message
            # fb_msg.send_msg(fname, lname, whatapp_contact_list)

        else:
            print("Unable to access DOB")


# -----------------------CSV File Reading----------------------

def read_csv():

    with open(dir_path + '/credentials/csv.csv') as f:
        reader = csv.reader(f, delimiter=',')

        for row in reader:
            try:
                if(row[4]==str(today)):
                    fname = row[0]
                    lname = row[1]
                    name = fname + " " + lname
                    mobile = row[2]
                    email = row[3]
                    whatapp_contact_list = []  # contacts to send message
                    for i in range(5, len(row)):
                        whatapp_name = row[i].split('_')

                        whatapp_name = '"' + ' '.join(whatapp_name) + '"'

                        whatapp_contact_list.append(whatapp_name)

                    print("Happy B'day " + name + " . Your mobile numbe is " + mobile + ". And your email is " + email )
                    # send_sms.sendPostRequest(str(mobile), fname)     #send sms
                    # send_mail.mail("rajp8340@gmail.com", fname)       #send mail
                    # whatsapp_msg.send_msg(whatapp_contact_list, fname)     #whatsapp message
                    # fb_msg.send_msg(fname, lname, whatapp_contact_list)
            except:
                print('Not Birthday')

# ---------------------Excel Reading----------------------

def read_excel():
    wb = load_workbook(dir_path + '/credentials/excel.xlsx')
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if(row[4]==str(today)):
                fname = row[0]
                lname = row[1]
                name = fname + " " + lname
                mobile = row[2]
                email = row[3]
                whatapp_contact_list = []  # contacts to send message
                for i in range(5, len(row)):
                    whatapp_name = row[i].split('_')

                    whatapp_name = '"' + ' '.join(whatapp_name) + '"'

                    whatapp_contact_list.append(whatapp_name)

                print("Happy B'day " + name + " . Your mobile numbe is " + mobile + ". And your email is " + email)
                # send_sms.sendPostRequest(str(mobile), fname)     #send sms
                # send_mail.mail("rajp8340@gmail.com", fname)       #send mail
                # whatsapp_msg.send_msg(whatapp_contact_list, fname)     #whatsapp message
                # fb_msg.send_msg(fname, lname, whatapp_contact_list)
        except:
            print('Not Birthday')


if os.path.exists(dir_path + '/log.txt'):
    if not log.log_read():
        log.log_write()
        read_txt()  #method to read data
else:
    log.log_write()
    read_txt()      #method to read data