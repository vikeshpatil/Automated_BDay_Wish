from log import *
import send_sms
import send_mail
import whatsapp_msg
import fb_msg

import time

import datetime
from openpyxl import load_workbook

import os
dir_path = os.path.dirname(os.path.realpath(__file__))

import csv

today = datetime.date.today()

#------------- Reading data in text file----------------------#
def read_txt():

    with open(dir_path + '/credentials/text.txt', mode='r') as f:
        readContent = f.readlines()
    for j in range(len(readContent)):
        if(readContent[j].split()[4] == str(today)):
            Person = readContent[j].split()
            fname = Person[0]
            lname = Person[1]
            name = Person[0] + " " + Person[1]
            mobile = Person[2]
            email = Person[3]
            whatapp_contact_list = []  # contacts to send message
            for i in range(5, len(Person)):
                whatapp_name = Person[i].split('_')

                whatapp_name = '"' + ' '.join(whatapp_name) + '"'

                whatapp_contact_list.append(whatapp_name)

            print("It's " + fname + "'s Happy B'day ")

            try:
                print('Sending SMS ...')
                send_sms.sendPostRequest(str(mobile), fname)  # send sms
                print('Sms sent.')
            except:
                print('Unable to send Sms.')

            try:
                print('Sending mail ...')
                send_mail.mail(email, fname)  # send mail
                print('Mail sent.')
            except:
                print('Unable to send mail')

            try:
                print('Sending wish on facebook ...')
                fb_msg.send_fmsg(fname, lname, whatapp_contact_list)  # facebook message
                print('Facebook message sent.')
            except:
                print('Unable to send Facebook message')

            try:
                print('Sending wish on whatsapp ...')
                whatsapp_msg.send_wmsg(whatapp_contact_list, fname)  # whatsapp message
                print('Whatsapp message sent.')
            except:
                print('Unable to send Whatsapp message')

        else:
            print('No one have birthday today :(')


# -----------------------CSV File Reading----------------------

def read_csv():

    with open(dir_path + '/credentials/csv.csv') as f:
        reader = csv.reader(f, delimiter=',')

        for row in reader:
            try:
                if(row[4]==str(today)):
                    fname = row[0]
                    lname = row[1]
                    name = fname + " " + lname
                    mobile = row[2]
                    email = row[3]
                    whatapp_contact_list = []  # contacts to send message
                    for i in range(5, len(row)):
                        whatapp_name = row[i].split('_')

                        whatapp_name = '"' + ' '.join(whatapp_name) + '"'

                        whatapp_contact_list.append(whatapp_name)

                    print("It's " + fname + "'s Happy B'day ")

                    try:
                        print('Sending SMS ...')
                        send_sms.sendPostRequest(str(mobile), fname)  # send sms
                        print('Sms sent.')
                    except:
                        print('Unable to send Sms.')

                    try:
                        print('Sending mail ...')
                        send_mail.mail(email, fname)  # send mail
                        print('Mail sent.')
                    except:
                        print('Unable to send mail')

                    try:
                        print('Sending wish on facebook ...')
                        fb_msg.send_fmsg(fname, lname, whatapp_contact_list)  # facebook message
                        print('Facebook message sent.')
                    except:
                        print('Unable to send Facebook message')

                    try:
                        print('Sending wish on whatsapp ...')
                        whatsapp_msg.send_wmsg(whatapp_contact_list, fname)  # whatsapp message
                        print('Whatsapp message sent.')
                    except:
                        print('Unable to send Whatsapp message')
            except:
                print('No one have birthday today :(')

# ---------------------Excel Reading----------------------
def read_excel():
    wb = load_workbook(dir_path + '/credentials/excel.xlsx')
    ws = wb.active
    print('Reading excel')
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:

            if(row[4]==str(today)):
                fname = row[0]
                lname = row[1]
                name = fname + " " + lname
                mobile = row[2]
                email = row[3]
                whatapp_contact_list = []  # contacts to send message
                for i in range(5, len(row)):
                    whatapp_name = row[i].split('_')

                    whatapp_name = '"' + ' '.join(whatapp_name) + '"'

                    whatapp_contact_list.append(whatapp_name)

                print("It's " + fname + "'s Happy B'day ")

                try:
                    print('Sending SMS ...')
                    send_sms.sendPostRequest(str(mobile), fname)  # send sms
                    print('Sms sent.')
                except:
                    print('Unable to send Sms.')

                try:
                    print('Sending mail ...')
                    send_mail.mail(email, fname)  # send mail
                    print('Mail sent.')
                except:
                    print('Unable to send mail')

                try:
                    print('Sending wish on facebook ...')
                    fb_msg.send_fmsg(fname, lname, whatapp_contact_list)  # facebook message
                    print('Facebook message sent.')
                except:
                    print('Unable to send Facebook message')

                try:
                    print('Sending wish on whatsapp ...')
                    whatsapp_msg.send_wmsg(whatapp_contact_list, fname)  # whatsapp message
                    print('Whatsapp message sent.')
                except:
                    print('Unable to send Whatsapp message')

        except:
            print('Done Scanning Birthdays!')

time.sleep(180)

if os.path.exists(dir_path + '/log.txt'):
    if not log_read():
        log_write()
        read_excel()  #method to read data
else:
    log_write()
    read_excel()      #method to read data
