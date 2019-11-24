
# import openpyxl and tkinter modules
from openpyxl import load_workbook, Workbook
from tkinter import *
import csv
import os

dir_path = os.path.dirname(os.path.realpath(__file__))

if not os.path.exists(dir_path + '/credentials/'):
    os.mkdir(dir_path + '/credentials')

try:
    wb = load_workbook(dir_path + '/credentials/excel.xlsx')
except:
    wb = Workbook()

sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "First Name"
    sheet.cell(row=1, column=2).value = "Last Name"
    sheet.cell(row=1, column=3).value = "Mobile Number"
    sheet.cell(row=1, column=4).value = "Email Id"
    sheet.cell(row=1, column=5).value = "Date Of Birth"
    sheet.cell(row=1, column=6).value = "Whatsapp Group"
    sheet.cell(row=1, column=7).value = "Facebook Group"


# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    lname_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    mobile_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    email_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
    dob_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    wgroup_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    fgroup_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    fname_field.delete(0, END)
    lname_field.delete(0, END)
    mobile_field.delete(0, END)
    email_field.delete(0, END)
    dob_field.delete(0, END)
    wgroup_field.delete(0, END)
    fgroup_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
    global esave_pop

    # if user not fill any entry
    # then print "empty input"
    if (fname_field.get() == "" and
            lname_field.get() == "" and
            mobile_field.get() == "" and
            email_field.get() == "" and
            dob_field.get() == "" and
            wgroup_field.get() == "" and
            fgroup_field.get() == ""):

        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = fname_field.get()
        sheet.cell(row=current_row + 1, column=2).value = lname_field.get()
        sheet.cell(row=current_row + 1, column=3).value = mobile_field.get()
        sheet.cell(row=current_row + 1, column=4).value = email_field.get()
        sheet.cell(row=current_row + 1, column=5).value = dob_field.get()
        sheet.cell(row=current_row + 1, column=6).value = wgroup_field.get()
        sheet.cell(row=current_row + 1, column=7).value = fgroup_field.get()

        # save the file
        wb.save(dir_path + '/credentials/excel.xlsx')

        esave_pop = Label(root, text="Details Saved in Excel File Successfully.", fg="Black", bg="light green", font=("calibri", 13))
        esave_pop.grid(row=13, column=1)


    #     ---------------Text File-------------------

def save_txt():

    global tsave_pop

    with open(dir_path + '/credentials/text.txt', 'a') as f:

        wgroup_name = str(wgroup_field.get()).split()
        wgroup_name = '_'.join(wgroup_name)
        fgroup_name = str(fgroup_field.get()).split()
        fgroup_name = '_'.join(fgroup_name)

        writecontent = f.write(str(fname_field.get()) + " " + str(lname_field.get()) + " " + str(mobile_field.get())
                               + " " + str(email_field.get()) + " " + str(dob_field.get()) + " "
                               + wgroup_name + " " + fgroup_name + "\n")

    tsave_pop = Label(root, text="Details Saved in Text File Successfully.", fg="Black", bg="light green", font=("calibri", 13))
    tsave_pop.grid(row=14, column=1)
    # Driver code

# -----------------------CSV File-----------------------
def save_csv():
    global csave_pop
    with open(dir_path + '/credentials/csv.csv', mode='a') as f:
        # fieldNames = ['First Name', 'Last Name', 'Mobile Number', 'Email', 'Date Of Birth', 'Whatsapp Group', 'Facebook Group']

        wgroup_name = str(wgroup_field.get()).split()
        wgroup_name = '_'.join(wgroup_name)
        fgroup_name = str(fgroup_field.get()).split()
        fgroup_name = '_'.join(fgroup_name)

        writer = csv.writer(f, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)

        writer.writerow([fname_field.get(), lname_field.get(), mobile_field.get(), email_field.get(), dob_field.get(), wgroup_name, fgroup_name])

    csave_pop = Label(root, text="Details Saved in CSV File Successfully.", fg="Black", bg="light green",
                      font=("calibri", 13))
    csave_pop.grid(row=15, column=1)


def Clear_Field():

    fname_field.focus_set()
    clear()
    csave_pop.destroy()
    esave_pop.destroy()
    tsave_pop.destroy()

if __name__ == "__main__":
    global root
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light green')

    # set the title of GUI window
    root.title("Contact Information")

    # set the configuration of GUI window
    root.geometry("700x500")

    excel()

    heading = Label(root, text="Fill the information for contacts", bg="light green", fg="black", width=30, height=2, font=("Calibri", 15))

    fname = Label(root, text="First Name", fg="black", bg="light green")

    lname = Label(root, text="Last Name", fg="black", bg="light green")

    mobile = Label(root, text="Mobile Number", fg="black", bg="light green")

    email = Label(root, text="Email", fg="black", bg="light green")

    dob= Label(root, text="Date Of Birth(yyyy-mm-dd)", fg="black", bg="light green")

    wgroup = Label(root, text="Whatsapp Group Name", fg="black", bg="light green")

    fgroup = Label(root, text="Facebook Group Name", fg="black", bg="light green")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    fname.grid(row=1, column=0)
    lname.grid(row=2, column=0)
    mobile.grid(row=3, column=0)
    email.grid(row=4, column=0)
    dob.grid(row=5, column=0)
    wgroup.grid(row=6, column=0)
    fgroup.grid(row=7, column=0)

    # create a text entry box
    # for typing the information
    fname_field = Entry(root)
    lname_field = Entry(root)
    mobile_field = Entry(root)
    email_field = Entry(root)
    dob_field = Entry(root)
    wgroup_field = Entry(root)
    fgroup_field = Entry(root)

    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    lname_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    mobile.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    email.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    dob.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    wgroup.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    fgroup.bind("<Return>", focus6)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    fname_field.grid(row=1, column=1, ipadx="100")
    lname_field.grid(row=2, column=1, ipadx="100")
    mobile_field.grid(row=3, column=1, ipadx="100")
    email_field.grid(row=4, column=1, ipadx="100")
    dob_field.grid(row=5, column=1, ipadx="100")
    wgroup_field.grid(row=6, column=1, ipadx="100")
    fgroup_field.grid(row=7, column=1, ipadx="100")

    # call excel function
    excel()

    save = Label(root, text="", bg="light green")
    save.grid(row=8, column=1)

    # create a Submit Button and place into the root window
    Esave = Button(root, text="Save Excel", fg="Black", width=10, height=2, bg="Red", command=insert)
    Esave.grid(row=9, column=1)

    Tsave = Button(root, text="Save Text File", fg="Black", width=10, height=2, bg="Red", command=save_txt)
    Tsave.grid(row=10, column=1)

    Csave = Button(root, text="Save CSV File", fg="Black", width=10, height=2, bg="Red", command=save_csv)
    Csave.grid(row=11, column=1)

    Clear = Button(root, text="Clear Fields", fg="Black", width=10, height=2, bg="Red", command=Clear_Field)
    Clear.grid(row=12, column=1)



    # start the GUI
    root.mainloop()